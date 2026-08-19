#!/usr/bin/env python3
"""Gera seeds SQL a partir da ficha cadastral Mandotti (Excel).

Uso:
  python3 scripts/import_ficha_cadastral.py [--xlsx PATH] [--out-dir supabase]

Saídas:
  supabase/seed_fazendas.sql          (atualizado)
  supabase/seed_maquinarios.sql
  supabase/seed_producao.sql
  supabase/seed_resumo_patrimonial.sql
  supabase/seed_patrimonio_bens.sql
  supabase/seed_grupo_contatos.sql
  supabase/seed_produtividade_institucional.sql
  supabase/seed_config_grupo.sql
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

DEFAULT_XLSX = (
    Path.home()
    / ".cursor/projects/workspace/uploads/FICHA_CADASTRAL_MANDOTTI__LIMPA__2__1b30.xlsx"
)

EMISSOR_EDER = "a0000000-0000-4000-8000-000000000001"
EMISSOR_NAGYLA = "a0000000-0000-4000-8000-000000000002"

FAZENDA_IDS: dict[str, str] = {
    "Cana Brava": "b0000001-0000-4000-8000-000000000001",
    "Cana Brava Abertura": "b0000001-0000-4000-8000-000000000013",
    "São José do Ribamar": "b0000001-0000-4000-8000-000000000002",
    "Pau Ferrado": "b0000001-0000-4000-8000-000000000003",
    "São Judas": "b0000001-0000-4000-8000-000000000004",
    "São Bento": "b0000001-0000-4000-8000-000000000005",
    "Salinas": "b0000001-0000-4000-8000-000000000006",
    "Cana Brava 2": "b0000001-0000-4000-8000-000000000007",
    "Brejão": "b0000001-0000-4000-8000-000000000008",
    "Barracão e São José": "b0000001-0000-4000-8000-000000000009",
    "Telha": "b0000001-0000-4000-8000-000000000010",
    "Sol Nascente": "b0000001-0000-4000-8000-000000000011",
    "Cruz de Malta": "b0000001-0000-4000-8000-000000000012",
}

MATRICULAS: dict[str, str | None] = {
    "São Judas": "2072",
    "Pau Ferrado": "2417",
    "São Bento": "167",
    "Salinas": None,
    "Brejão": "6614",
    "São José do Ribamar": "3843",
    "Cana Brava": "8714",
    "Barracão e São José": "1110",
    "Telha": None,
}


def sql_str(value: str | None) -> str:
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''").strip() + "'"


def sql_num(value: float | int | None) -> str:
    if value is None:
        return "null"
    return str(float(value))


def sql_bool(value: bool) -> str:
    return "true" if value else "false"


def strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(c for c in normalized if not unicodedata.combining(c))


def norm_fazenda(name: str | None) -> str | None:
    if not name:
        return None
    n = strip_accents(str(name).upper().strip())
    n = re.sub(r"^FAZENDA\s+", "", n)
    n = re.sub(r"^FAZ\.\s*", "", n)
    n = re.sub(r"\s+", " ", n)
    mapping = {
        "CANA BRAVA": "Cana Brava",
        "CANA BRAVA ABERTURA": "Cana Brava Abertura",
        "CANA BRAVA 2  ABERTURA": "Cana Brava 2",
        "CANA BRAVA 2 ABERTURA": "Cana Brava 2",
        "PAU FERRADO": "Pau Ferrado",
        "SAO JUDAS": "São Judas",
        "SÃO JUDAS": "São Judas",
        "SAO BENTO": "São Bento",
        "SÃO BENTO": "São Bento",
        "SALINAS": "Salinas",
        "SAO JOSE DO RIBAMAR": "São José do Ribamar",
        "SÃO JOSE DO RIBAMAR": "São José do Ribamar",
        "BREJAO": "Brejão",
        "BARRACAO": "Barracão e São José",
        "BARR E SAO JOSE": "Barracão e São José",
        "TELHA": "Telha",
        "SOLNASCENTE": "Sol Nascente",
        "CRUZ DEMALTA": "Cruz de Malta",
    }
    for key, val in mapping.items():
        if key in n:
            return val
    return str(name).strip()


def norm_cultura(raw: str | None) -> str | None:
    if not raw:
        return None
    c = strip_accents(str(raw).upper().strip())
    if "SOJA" in c:
        return "soja"
    if "MILHO" in c:
        return "milho"
    if "SORGO" in c:
        return "sorgo"
    if "MILHETO" in c:
        return "milheto"
    return None


def norm_safra(raw) -> str:
    if isinstance(raw, (int, float)):
        year = int(raw)
        if year >= 2023:
            return f"{year}/{str(year + 1)[-2:]}"
        return str(year)
    s = str(raw).strip()
    s = s.replace("2028/2029", "2028/29").replace("2029/2030", "2029/30")
    m = re.match(r"^(\d{4})/(\d{2,4})$", s)
    if m:
        y1, y2 = m.group(1), m.group(2)
        return f"{y1}/{y2[-2:]}"
    if re.match(r"^\d{4}$", s):
        y = int(s)
        return f"{y}/{str(y + 1)[-2:]}"
    return s


def parse_custo(raw) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).replace(".", "").replace(",", ".")
    m = re.search(r"([\d.]+)", s)
    return float(m.group(1)) if m else None


def emissor_from_proprietario(nome: str | None) -> str:
    if not nome:
        return EMISSOR_EDER
    n = strip_accents(str(nome).upper())
    if "NAGYLA" in n:
        return EMISSOR_NAGYLA
    return EMISSOR_EDER


def parse_maquinarios(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Maquinários"]
    items: list[dict] = []
    for row in range(4, ws.max_row + 1):
        ordem_raw = ws.cell(row, 2).value
        if ordem_raw is None:
            continue
        if str(ordem_raw).upper() == "TOTAL":
            break
        ordem = int(float(ordem_raw))
        nome = ws.cell(row, 3).value
        categoria = ws.cell(row, 4).value or "Outros"
        proprietario = ws.cell(row, 5).value
        fazenda_nome = ws.cell(row, 6).value
        marca = ws.cell(row, 7).value
        modelo = ws.cell(row, 8).value
        ano_raw = ws.cell(row, 9).value
        valor = ws.cell(row, 10).value
        cor = ws.cell(row, 11).value
        chassi = ws.cell(row, 12).value

        if not nome and modelo:
            nome = str(modelo).strip()
        if not nome:
            nome = f"Item {ordem}"

        ano = int(float(ano_raw)) if ano_raw is not None else None
        fazenda_key = norm_fazenda(str(fazenda_nome) if fazenda_nome else None)

        items.append(
            {
                "ordem": ordem,
                "nome": str(nome).strip(),
                "categoria": str(categoria).strip(),
                "emissor_id": emissor_from_proprietario(
                    str(proprietario) if proprietario else None
                ),
                "fazenda_id": FAZENDA_IDS.get(fazenda_key) if fazenda_key else None,
                "fazenda_nome": fazenda_key or (str(fazenda_nome).strip() if fazenda_nome else None),
                "marca": str(marca).strip() if marca else None,
                "modelo": str(modelo).strip() if modelo else None,
                "ano": ano,
                "valor_aquisicao": float(valor) if valor is not None else None,
                "cor": str(cor).strip() if cor else None,
                "chassi_serie": str(chassi).strip() if chassi else None,
            }
        )
    return items


def parse_producao_grupo(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Dados Operacionais"]
    rows: list[dict] = []
    for row in range(10, 35):
        safra_raw = ws.cell(row, 2).value
        cultura_raw = ws.cell(row, 5).value or ws.cell(row, 4).value
        if not safra_raw or not cultura_raw:
            continue
        if "FONTES" in str(safra_raw).upper():
            break
        cultura_label = str(cultura_raw).strip().upper()
        cultura = norm_cultura(cultura_label)
        if not cultura:
            continue
        ciclo = "safrinha" if "SAFRA" in cultura_label and cultura == "milho" else "safra"
        area = ws.cell(row, 14).value
        prod = ws.cell(row, 18).value
        preco = ws.cell(row, 21).value
        custo = parse_custo(ws.cell(row, 24).value)
        if area is None and prod is None:
            continue
        safra = norm_safra(safra_raw)
        tipo = "realizado" if prod is not None and preco is not None else "projecao"
        rows.append(
            {
                "safra": safra,
                "cultura_codigo": cultura,
                "area_plantio_ha": float(area) if area is not None else None,
                "produtividade_sc_ha": float(prod) if prod is not None else None,
                "preco_saca": float(preco) if preco is not None else None,
                "custo_saca": custo,
                "tipo": tipo,
                "ciclo": ciclo,
            }
        )
    return rows


def parse_producao_fazenda(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Dados Operacionais"]
    rows: list[dict] = []
    for row in range(41, 80):
        safra_raw = ws.cell(row, 2).value
        cultura_raw = ws.cell(row, 4).value
        imovel = ws.cell(row, 7).value
        if not safra_raw or not cultura_raw or not imovel:
            continue
        if "DESPESA" in str(safra_raw).upper():
            break
        cultura = norm_cultura(str(cultura_raw))
        fazenda = norm_fazenda(str(imovel))
        if not cultura or not fazenda:
            continue
        matricula = ws.cell(row, 13).value
        area = ws.cell(row, 15).value
        prod = ws.cell(row, 18).value
        rows.append(
            {
                "fazenda_id": FAZENDA_IDS.get(fazenda),
                "safra": norm_safra(safra_raw),
                "cultura_codigo": cultura,
                "area_plantio_ha": float(area) if area is not None else None,
                "produtividade_sc_ha": float(prod) if prod is not None else None,
                "matricula": str(int(matricula)) if isinstance(matricula, (int, float)) else None,
            }
        )
    return rows


def parse_resumo(wb: openpyxl.Workbook) -> dict:
    ws = wb["Resumo Executivo"]
    return {
        "participacoes_societarias": ws.cell(4, 2).value or 0,
        "imoveis": ws.cell(5, 2).value or 0,
        "maquinarios_veiculos": ws.cell(6, 2).value or 0,
        "animais": ws.cell(7, 2).value or 0,
        "outros_bens": ws.cell(8, 2).value or 0,
        "patrimonio_total": ws.cell(9, 2).value or 0,
        "passivo_total": ws.cell(10, 2).value or 0,
        "patrimonio_liquido": ws.cell(11, 2).value or 0,
        "endividamento_pct": ws.cell(12, 2).value or 0,
        "total_projetado_juros": ws.cell(13, 2).value,
        "passivo_eder": ws.cell(6, 5).value,
        "passivo_nagyla": ws.cell(7, 5).value,
        "cronograma_ate_jun26": ws.cell(6, 8).value,
        "cronograma_jul26_jun27": ws.cell(6, 9).value,
        "cronograma_jul27_jun28": ws.cell(6, 10).value,
        "cronograma_jul28_jun29": ws.cell(6, 11).value,
        "cronograma_jul29_jun30": ws.cell(6, 12).value,
        "cronograma_apos_jun30": ws.cell(6, 13).value,
    }


def parse_passivo_instituicao(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Resumo Executivo"]
    rows = []
    for r in range(12, 20):
        inst = ws.cell(r, 4).value
        saldo = ws.cell(r, 5).value
        if not inst or not saldo:
            continue
        if "critério" in str(inst).lower():
            continue
        rows.append({"instituicao": str(inst).strip(), "saldo_devedor": float(saldo)})
    return rows


def parse_patrimonio_bens(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Dados Patrimoniais"]
    rows: list[dict] = []
    mode = None
    ordem = 0
    for r in range(6, 45):
        b = ws.cell(r, 2).value
        if not b:
            continue
        text = str(b).strip()
        upper = strip_accents(text.upper())
        if "PARTICIPA" in upper and "EMPRESA" in upper:
            mode = "participacao"
            continue
        if "BENS IM" in upper:
            mode = "imovel"
            continue
        if "BENFEITORIA" in upper or "MAQUINAS / VEICULOS" in upper:
            break
        if text.startswith("  ") and "-" in text[:6]:
            continue
        if mode == "participacao":
            ordem += 1
            rows.append({"tipo": "participacao", "descricao": text, "municipio": None, "ordem": ordem})
        elif mode == "imovel":
            municipio = ws.cell(r, 8).value
            muni = str(municipio).strip() if municipio else None
            b3 = ws.cell(r, 3).value
            if isinstance(b, (int, float)) and b3:
                desc = str(b3).strip()
            else:
                desc = text
            if desc in ("1-", "1.0") or len(desc) <= 2:
                continue
            ordem += 1
            rows.append({"tipo": "imovel", "descricao": desc, "municipio": muni, "ordem": ordem})
    return rows


def parse_grupo_contatos(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Dados Cadastrais"]
    rows: list[dict] = []
    mode = None
    ordem = 0

    def add(cat: str, nome: str, cidade=None, contato=None, agencia=None):
        nonlocal ordem
        ordem += 1
        ag = str(int(agencia)) if isinstance(agencia, (int, float)) else agencia
        rows.append(
            {
                "categoria": cat,
                "nome": nome.strip(),
                "cidade": str(cidade).strip() if cidade else None,
                "contato_nome": str(contato).strip() if contato else None,
                "agencia": str(ag).strip() if ag else None,
                "ordem": ordem,
            }
        )

    for r in range(36, 70):
        b = ws.cell(r, 2).value
        if not b:
            continue
        text = str(b).strip()
        upper = strip_accents(text.upper())
        if "FONTES DE REFER" in upper:
            mode = "ref"
            continue
        if "PRINCIPAIS FORNECEDORES" in upper:
            mode = "fornecedor"
            continue
        if "DESTINO DA PRODU" in upper:
            mode = "destino"
            continue
        if "CONDOM" in upper:
            break
        if text.startswith("  ") or text in ("NOME DA EMPRESA", "BANCÁRIA : NOME DO BANCO", "RUA, AVENIDA, TRAVESSA, ETC"):
            continue
        if mode == "ref":
            if upper == "COMERCIAL":
                nxt = ws.cell(r + 1, 2).value
                if nxt:
                    add("referencia_comercial", str(nxt))
                continue
            if upper == "PESSOAL":
                nxt = ws.cell(r + 1, 2).value
                if nxt:
                    add("referencia_pessoal", str(nxt))
                continue
            if "BANC" in upper:
                banco = text.split(":")[-1].strip() if ":" in text else text
                contato = ws.cell(r, 8).value
                agencia = ws.cell(r, 12).value
                add("referencia_bancaria", banco or "Banco", contato=contato, agencia=agencia)
        elif mode == "fornecedor":
            cidade = ws.cell(r, 11).value
            if "NOME" not in upper:
                add("fornecedor", text, cidade=cidade)
        elif mode == "destino":
            cidade = ws.cell(r, 11).value
            if "NOME" not in upper:
                add("destino_producao", text, cidade=cidade)
    return rows


def parse_institucional(wb: openpyxl.Workbook) -> tuple[dict, list[dict]]:
    ws = wb["Institucional"]
    config = {
        "tempo_agricultura_anos": 25,
        "perfil_grupo": None,
        "fonte_bancario_pct": None,
        "fonte_proprio_pct": None,
        "fonte_principal_banco": None,
    }
    perfil = ws.cell(8, 2).value
    if perfil:
        config["perfil_grupo"] = str(perfil).replace("\n", " ").strip()[:2000]
    tempo = ws.cell(6, 2).value
    if tempo and "25" in str(tempo):
        config["tempo_agricultura_anos"] = 25

    ws_op = wb["Dados Operacionais"]
    config["fonte_bancario_pct"] = ws_op.cell(34, 5).value
    config["fonte_proprio_pct"] = ws_op.cell(34, 10).value
    config["fonte_principal_banco"] = ws_op.cell(34, 17).value
    if config["fonte_principal_banco"]:
        config["fonte_principal_banco"] = str(config["fonte_principal_banco"]).strip()

    benchmarks = []
    for r in range(25, 28):
        safra_soja = ws.cell(r, 2).value
        prod_soja = ws.cell(r, 3).value
        safra_milho = ws.cell(r, 5).value
        prod_milho = ws.cell(r, 6).value
        if safra_soja and prod_soja:
            benchmarks.append(
                {"safra": norm_safra(safra_soja), "cultura_codigo": "soja", "produtividade_sc_ha": float(prod_soja)}
            )
        if safra_milho and prod_milho:
            benchmarks.append(
                {"safra": norm_safra(safra_milho), "cultura_codigo": "milho", "produtividade_sc_ha": float(prod_milho)}
            )
    return config, benchmarks


def write_fazendas_seed(out_dir: Path) -> None:
    lines = [
        "-- Carga: aba Institucional + matrículas (Dados Operacionais)",
        "-- Gerado por scripts/import_ficha_cadastral.py",
        "",
        "insert into public.fazendas (",
        "  id, emissor_id, nome, codigo, municipio, uf, regime,",
        "  area_produtiva_ha, area_abertura_ha, area_hectares, matricula,",
        "  venc_arrendamento, inclui_quadro_produtivo, custo_arrendamento, observacoes, ativo",
        ") values",
        f"  ('{FAZENDA_IDS['Cana Brava']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'Cana Brava', 'CB-01', 'Pedro Afonso', 'TO', 'arrendada',",
        "   1430, null, 1430, '8714', '2044-12-31', true, null,",
        "   'Planilha: 1430 ha produtivos. Culturas 26/27: Soja → Sorgo', true),",
        f"  ('{FAZENDA_IDS['Cana Brava Abertura']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'Cana Brava Abertura', 'CB-AB', 'Pedro Afonso', 'TO', 'arrendada',",
        "   null, 1000, 1000, null, '2044-12-31', false, null,",
        "   'Área em abertura/licenciamento — linha separada na planilha', true),",
        f"  ('{FAZENDA_IDS['São José do Ribamar']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'São José do Ribamar', 'SJR-01', 'Santa Maria do Tocantins', 'TO', 'arrendada',",
        "   1700, null, 1700, '3843', '2037-12-31', true, null,",
        "   'Culturas 26/27: Soja → Milho', true),",
        f"  ('{FAZENDA_IDS['Pau Ferrado']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'Pau Ferrado', 'PF-01', 'Pedro Afonso', 'TO', 'arrendada',",
        "   450, null, 450, '2417', '2032-12-31', true, null,",
        "   'Culturas 26/27: Soja → Milho', true),",
        f"  ('{FAZENDA_IDS['São Judas']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'São Judas', 'SJ-01', 'Pedro Afonso', 'TO', 'arrendada',",
        "   450, null, 450, '2072', '2033-12-31', true, null,",
        "   'Culturas 26/27: Soja → Sorgo', true),",
        f"  ('{FAZENDA_IDS['São Bento']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'São Bento', 'SB-01', 'Santa Maria do Tocantins', 'TO', 'arrendada',",
        "   400, null, 400, '167', '2035-12-31', true, null,",
        "   'Culturas 26/27: Soja → Milho', true),",
        f"  ('{FAZENDA_IDS['Salinas']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'Salinas', 'SL-01', 'Pedro Afonso', 'TO', 'arrendada',",
        "   60, null, 60, null, '2032-12-31', true, null,",
        "   'Culturas 26/27: Soja → Sorgo', true),",
        f"  ('{FAZENDA_IDS['Cana Brava 2']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'Cana Brava 2', 'CB-02', 'Pedro Afonso', 'TO', 'arrendada',",
        "   null, 430, 430, null, '2036-12-31', false, null,",
        "   'Abertura 430 ha — fora do quadro produtivo atual', true),",
        f"  ('{FAZENDA_IDS['Brejão']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'Brejão', 'BR-01', 'Pedro Afonso', 'TO', 'propria',",
        "   70, null, 70, '6614', null, true, null, 'Própria', true),",
        f"  ('{FAZENDA_IDS['Barracão e São José']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'Barracão e São José', 'BSJ-01', 'Pedro Afonso', 'TO', 'propria',",
        "   700, null, 700, '1110', null, true, null, 'Própria', true),",
        f"  ('{FAZENDA_IDS['Telha']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'Telha', 'TL-01', 'Pedro Afonso', 'TO', 'propria',",
        "   500, null, 500, null, null, true, null, 'Própria — milho safra', true),",
        f"  ('{FAZENDA_IDS['Sol Nascente']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'Sol Nascente', 'SN-01', 'Pedro Afonso', 'TO', 'arrendada_a_terceiro',",
        "   null, null, null, null, null, false, null,",
        "   'Arrendada a terceiros (gado) — fora do quadro produtivo', true),",
        f"  ('{FAZENDA_IDS['Cruz de Malta']}'::uuid, '{EMISSOR_EDER}'::uuid,",
        "   'Cruz de Malta', 'CM-01', 'Pedro Afonso', 'TO', 'arrendada_a_terceiro',",
        "   null, null, null, null, null, false, null,",
        "   'Arrendada a terceiros (gado) — fora do quadro produtivo', true)",
        "on conflict (id) do update set",
        "  nome = excluded.nome,",
        "  codigo = excluded.codigo,",
        "  municipio = excluded.municipio,",
        "  uf = excluded.uf,",
        "  regime = excluded.regime,",
        "  area_produtiva_ha = excluded.area_produtiva_ha,",
        "  area_abertura_ha = excluded.area_abertura_ha,",
        "  area_hectares = excluded.area_hectares,",
        "  matricula = excluded.matricula,",
        "  venc_arrendamento = excluded.venc_arrendamento,",
        "  inclui_quadro_produtivo = excluded.inclui_quadro_produtivo,",
        "  observacoes = excluded.observacoes,",
        "  updated_at = now();",
    ]
    (out_dir / "seed_fazendas.sql").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_maquinarios_seed(out_dir: Path, items: list[dict]) -> None:
    lines = [
        "-- Carga: aba Maquinários (78 itens + ordem 79 BAZUCA)",
        "-- Gerado por scripts/import_ficha_cadastral.py",
        "",
        "-- Remove registros antigos e recarrega por ordem (idempotente para dev)",
        "delete from public.maquinarios where origem_planilha = true;",
        "",
        "insert into public.maquinarios (",
        "  emissor_id, fazenda_id, nome, categoria, fazenda_nome, marca, modelo,",
        "  ano, valor_aquisicao, cor, chassi_serie, ordem, origem_planilha",
        ") values",
    ]
    # Need origem_planilha column - I'll add to migration or use a comment field
    # Actually I should not add origem_planilha if not in schema. Use upsert by ordem instead.

    values = []
    for item in items:
        values.append(
            "  ("
            + f"{sql_str(item['emissor_id'])}::uuid, "
            + f"{sql_str(item['fazenda_id']) + '::uuid' if item['fazenda_id'] else 'null'}, "
            + f"{sql_str(item['nome'])}, "
            + f"{sql_str(item['categoria'])}, "
            + f"{sql_str(item['fazenda_nome'])}, "
            + f"{sql_str(item['marca'])}, "
            + f"{sql_str(item['modelo'])}, "
            + f"{sql_num(item['ano'])}, "
            + f"{sql_num(item['valor_aquisicao'])}, "
            + f"{sql_str(item['cor'])}, "
            + f"{sql_str(item['chassi_serie'])}, "
            + f"{item['ordem']}"
            + ")"
        )

    lines = [
        "-- Carga: aba Maquinários",
        "-- Gerado por scripts/import_ficha_cadastral.py",
        "",
        "-- Upsert por ordem (chave natural da planilha)",
        "insert into public.maquinarios (",
        "  emissor_id, fazenda_id, nome, categoria, fazenda_nome, marca, modelo,",
        "  ano, valor_aquisicao, cor, chassi_serie, ordem",
        ") values",
        ",\n".join(values),
        "on conflict (ordem) where deleted_at is null do update set",
        "  emissor_id = excluded.emissor_id,",
        "  fazenda_id = excluded.fazenda_id,",
        "  nome = excluded.nome,",
        "  categoria = excluded.categoria,",
        "  fazenda_nome = excluded.fazenda_nome,",
        "  marca = excluded.marca,",
        "  modelo = excluded.modelo,",
        "  ano = excluded.ano,",
        "  valor_aquisicao = excluded.valor_aquisicao,",
        "  cor = excluded.cor,",
        "  chassi_serie = excluded.chassi_serie,",
        "  updated_at = now();",
    ]
    (out_dir / "seed_maquinarios.sql").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_producao_seed(out_dir: Path, grupo: list[dict], fazenda: list[dict]) -> None:
    grupo_vals = []
    for r in grupo:
        grupo_vals.append(
            "  ("
            + f"{sql_str(r['safra'])}, "
            + f"{sql_str(r['cultura_codigo'])}, "
            + f"{sql_num(r['area_plantio_ha'])}, "
            + f"{sql_num(r['produtividade_sc_ha'])}, "
            + f"{sql_num(r['preco_saca'])}, "
            + f"{sql_num(r['custo_saca'])}, "
            + f"{sql_str(r['tipo'])}, "
            + f"{sql_str(r['ciclo'])}"
            + ")"
        )

    faz_vals = []
    for r in fazenda:
        faz_vals.append(
            "  ("
            + f"{sql_str(r['fazenda_id']) + '::uuid' if r['fazenda_id'] else 'null'}, "
            + f"{sql_str(r['safra'])}, "
            + f"{sql_str(r['cultura_codigo'])}, "
            + f"{sql_num(r['area_plantio_ha'])}, "
            + f"{sql_num(r['produtividade_sc_ha'])}, "
            + f"{sql_str(r['matricula'])}"
            + ")"
        )

    content = "\n".join(
        [
            "-- Carga: aba Dados Operacionais",
            "-- Gerado por scripts/import_ficha_cadastral.py",
            "",
            "insert into public.producao_grupo_safra (",
            "  safra, cultura_codigo, area_plantio_ha, produtividade_sc_ha,",
            "  preco_saca, custo_saca, tipo, ciclo",
            ") values",
            ",\n".join(grupo_vals),
            "on conflict (safra, cultura_codigo, tipo, ciclo) do update set",
            "  area_plantio_ha = excluded.area_plantio_ha,",
            "  produtividade_sc_ha = excluded.produtividade_sc_ha,",
            "  preco_saca = excluded.preco_saca,",
            "  custo_saca = excluded.custo_saca,",
            "  updated_at = now();",
            "",
            "insert into public.producao_fazenda_safra (",
            "  fazenda_id, safra, cultura_codigo, area_plantio_ha,",
            "  produtividade_sc_ha, matricula",
            ") values",
            ",\n".join(faz_vals),
            "on conflict (fazenda_id, safra, cultura_codigo) do update set",
            "  area_plantio_ha = excluded.area_plantio_ha,",
            "  produtividade_sc_ha = excluded.produtividade_sc_ha,",
            "  matricula = excluded.matricula,",
            "  updated_at = now();",
        ]
    )
    (out_dir / "seed_producao.sql").write_text(content + "\n", encoding="utf-8")


def write_resumo_seed(out_dir: Path, resumo: dict) -> None:
    content = "\n".join(
        [
            "-- Carga: aba Resumo Executivo",
            "-- Gerado por scripts/import_ficha_cadastral.py",
            "",
            "insert into public.resumo_patrimonial (",
            "  id, participacoes_societarias, imoveis, maquinarios_veiculos, animais,",
            "  outros_bens, patrimonio_total, passivo_total, patrimonio_liquido,",
            "  endividamento_pct, total_projetado_juros, passivo_eder, passivo_nagyla, referencia,",
            "  cronograma_ate_jun26, cronograma_jul26_jun27, cronograma_jul27_jun28,",
            "  cronograma_jul28_jun29, cronograma_jul29_jun30, cronograma_apos_jun30",
            ") values (",
            "  'd0000000-0000-4000-8000-000000000001'::uuid,",
            f"  {sql_num(resumo['participacoes_societarias'])},",
            f"  {sql_num(resumo['imoveis'])},",
            f"  {sql_num(resumo['maquinarios_veiculos'])},",
            f"  {sql_num(resumo['animais'])},",
            f"  {sql_num(resumo['outros_bens'])},",
            f"  {sql_num(resumo['patrimonio_total'])},",
            f"  {sql_num(resumo['passivo_total'])},",
            f"  {sql_num(resumo['patrimonio_liquido'])},",
            f"  {sql_num(resumo['endividamento_pct'])},",
            f"  {sql_num(resumo['total_projetado_juros'])},",
            f"  {sql_num(resumo['passivo_eder'])},",
            f"  {sql_num(resumo['passivo_nagyla'])},",
            f"  '{date.today().isoformat()}'::date,",
            f"  {sql_num(resumo.get('cronograma_ate_jun26'))},",
            f"  {sql_num(resumo.get('cronograma_jul26_jun27'))},",
            f"  {sql_num(resumo.get('cronograma_jul27_jun28'))},",
            f"  {sql_num(resumo.get('cronograma_jul28_jun29'))},",
            f"  {sql_num(resumo.get('cronograma_jul29_jun30'))},",
            f"  {sql_num(resumo.get('cronograma_apos_jun30'))}",
            ")",
            "on conflict (id) do update set",
            "  participacoes_societarias = excluded.participacoes_societarias,",
            "  imoveis = excluded.imoveis,",
            "  maquinarios_veiculos = excluded.maquinarios_veiculos,",
            "  animais = excluded.animais,",
            "  outros_bens = excluded.outros_bens,",
            "  patrimonio_total = excluded.patrimonio_total,",
            "  passivo_total = excluded.passivo_total,",
            "  patrimonio_liquido = excluded.patrimonio_liquido,",
            "  endividamento_pct = excluded.endividamento_pct,",
            "  total_projetado_juros = excluded.total_projetado_juros,",
            "  passivo_eder = excluded.passivo_eder,",
            "  passivo_nagyla = excluded.passivo_nagyla,",
            "  referencia = excluded.referencia,",
            "  cronograma_ate_jun26 = excluded.cronograma_ate_jun26,",
            "  cronograma_jul26_jun27 = excluded.cronograma_jul26_jun27,",
            "  cronograma_jul27_jun28 = excluded.cronograma_jul27_jun28,",
            "  cronograma_jul28_jun29 = excluded.cronograma_jul28_jun29,",
            "  cronograma_jul29_jun30 = excluded.cronograma_jul29_jun30,",
            "  cronograma_apos_jun30 = excluded.cronograma_apos_jun30,",
            "  updated_at = now();",
        ]
    )
    (out_dir / "seed_resumo_patrimonial.sql").write_text(content + "\n", encoding="utf-8")


def _generic_upsert_seed(
    out_dir: Path,
    filename: str,
    table: str,
    columns: list[str],
    conflict: str,
    rows: list[dict],
) -> None:
    if not rows:
        return
    vals = []
    for row in rows:
        vals.append("  (" + ", ".join(sql_str(row.get(c)) if c != "ordem" else str(row.get(c, 0)) for c in columns) + ")")
    content = "\n".join(
        [
            f"-- Gerado por scripts/import_ficha_cadastral.py",
            "",
            f"insert into public.{table} ({', '.join(columns)}) values",
            ",\n".join(vals),
            f"on conflict {conflict} do update set",
            ",\n".join(f"  {c} = excluded.{c}" for c in columns if c not in conflict.strip("()").split(",")[0].split()[-1:]),
            "  updated_at = now();",
        ]
    )
    (out_dir / filename).write_text(content + "\n", encoding="utf-8")


def write_complement_seeds(
    out_dir: Path,
    patrimonio: list[dict],
    contatos: list[dict],
    benchmarks: list[dict],
    passivo_inst: list[dict],
    config: dict,
) -> None:
    if patrimonio:
        vals = []
        for r in patrimonio:
            vals.append(
                f"  ({sql_str(r['tipo'])}, {sql_str(r['descricao'])}, {sql_str(r['municipio'])}, {r['ordem']})"
            )
        (out_dir / "seed_patrimonio_bens.sql").write_text(
            "\n".join(
                [
                    "-- aba Dados Patrimoniais",
                    "delete from public.patrimonio_bens where origem = 'planilha';",
                    "insert into public.patrimonio_bens (tipo, descricao, municipio, ordem) values",
                    ",\n".join(vals),
                    ";",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

    if contatos:
        vals = []
        for r in contatos:
            vals.append(
                "  ("
                + f"{sql_str(r['categoria'])}, {sql_str(r['nome'])}, {sql_str(r['cidade'])}, "
                + f"{sql_str(r['contato_nome'])}, {sql_str(r['agencia'])}, {r['ordem']}"
                + ")"
            )
        (out_dir / "seed_grupo_contatos.sql").write_text(
            "\n".join(
                [
                    "-- aba Dados Cadastrais (referências, fornecedores, destinos)",
                    "delete from public.grupo_contatos where origem = 'planilha';",
                    "insert into public.grupo_contatos (categoria, nome, cidade, contato_nome, agencia, ordem) values",
                    ",\n".join(vals),
                    ";",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

    if benchmarks:
        vals = []
        for r in benchmarks:
            vals.append(
                f"  ({sql_str(r['safra'])}, {sql_str(r['cultura_codigo'])}, {sql_num(r['produtividade_sc_ha'])})"
            )
        (out_dir / "seed_produtividade_institucional.sql").write_text(
            "\n".join(
                [
                    "-- aba Institucional (benchmarks produtividade)",
                    "insert into public.produtividade_institucional (safra, cultura_codigo, produtividade_sc_ha) values",
                    ",\n".join(vals),
                    "on conflict (safra, cultura_codigo) do update set",
                    "  produtividade_sc_ha = excluded.produtividade_sc_ha, updated_at = now();",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

    if passivo_inst:
        vals = []
        for r in passivo_inst:
            vals.append(f"  ({sql_str(r['instituicao'])}, {sql_num(r['saldo_devedor'])})")
        extra = "\n".join(
            [
                "-- passivo por instituição (Resumo Executivo)",
                "insert into public.passivo_por_instituicao (instituicao, saldo_devedor) values",
                ",\n".join(vals),
                "on conflict (instituicao) do update set saldo_devedor = excluded.saldo_devedor, updated_at = now();",
            ]
        )
        (out_dir / "seed_resumo_patrimonial.sql").write_text(
            (out_dir / "seed_resumo_patrimonial.sql").read_text(encoding="utf-8") + "\n" + extra + "\n",
            encoding="utf-8",
        )

    (out_dir / "seed_config_grupo.sql").write_text(
        "\n".join(
            [
                "-- Institucional + fontes de recurso (Operacional)",
                "update public.configuracoes_grupo set",
                f"  tempo_agricultura_anos = {config.get('tempo_agricultura_anos') or 'null'},",
                f"  perfil_grupo = {sql_str(config.get('perfil_grupo'))},",
                f"  fonte_bancario_pct = {sql_num(config.get('fonte_bancario_pct'))},",
                f"  fonte_proprio_pct = {sql_num(config.get('fonte_proprio_pct'))},",
                f"  fonte_principal_banco = {sql_str(config.get('fonte_principal_banco'))},",
                "  updated_at = now()",
                "where id = 'c0000000-0000-4000-8000-000000000001'::uuid;",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa ficha cadastral Mandotti para SQL seeds")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out-dir", type=Path, default=Path("supabase"))
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Planilha não encontrada: {args.xlsx}")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    maquinarios = parse_maquinarios(wb)
    grupo = parse_producao_grupo(wb)
    fazenda = parse_producao_fazenda(wb)
    resumo = parse_resumo(wb)
    config, benchmarks = parse_institucional(wb)
    patrimonio = parse_patrimonio_bens(wb)
    contatos = parse_grupo_contatos(wb)
    passivo_inst = parse_passivo_instituicao(wb)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    write_fazendas_seed(args.out_dir)
    write_maquinarios_seed(args.out_dir, maquinarios)
    write_producao_seed(args.out_dir, grupo, fazenda)
    write_resumo_seed(args.out_dir, resumo)
    write_complement_seeds(args.out_dir, patrimonio, contatos, benchmarks, passivo_inst, config)

    print(
        f"OK — {len(maquinarios)} maquinários, {len(grupo)} safras, {len(fazenda)} fazendas, "
        f"{len(patrimonio)} bens, {len(contatos)} contatos, {len(benchmarks)} benchmarks"
    )
    print(f"Seeds em {args.out_dir.resolve()}")


if __name__ == "__main__":
    main()
